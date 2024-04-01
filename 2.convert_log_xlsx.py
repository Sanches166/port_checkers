import json
import re
from openpyxl import Workbook

# создаем новый файл Excel и выбираем активный лист
wb = Workbook()
ws = wb.active

# задаем заголовки для столбцов
ws['A1'] = 'Protocol'
ws['B1'] = 'Source IP'
ws['C1'] = 'Source Port'
ws['D1'] = 'Destination IP'
ws['E1'] = 'Destination Port'
ws['F1'] = 'Flags'

# открываем файл с логами и читаем его построчно
with open('connections.log') as f:
    logs = f.readlines()

# начинаем запись данных со второй строки (первая строка занята заголовками)
row = 2

for log in logs:
    # используем регулярное выражение для поиска json в строке лога
    match = re.search("{.*\('(.+)', (\d+).+?(\d+)}", log)
    if match:
        # извлекаем найденный json и десериализуем его
        print(match.group())
        data = json.loads(match.group())
        # записываем данные в соответствующие ячейки в файле Excel
        ws.cell(row=row, column=1, value=data['proto'])
        ws.cell(row=row, column=2, value=data['src']['ip'])
        ws.cell(row=row, column=3, value=data['src']['port'])
        ws.cell(row=row, column=4, value=data['dst']['ip'])
        ws.cell(row=row, column=5, value=data['dst']['port'])
        if 'flags' in data:
            ws.cell(row=row, column=6, value=', '.join(data['flags']))
        # увеличиваем значение счетчика для перехода на следующую строку в файле Excel
        row += 1

# сохраняем файл Excel
wb.save('logs4.xlsx')